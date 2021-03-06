from openpyxl import load_workbook
from random import randint
import datetime


def isProjectPeriod(hr,day):
    projectperiodnamegiveninclasstt = ["PROJECT"]
    ret = False
    if( classtt.cell(column=classH[hr], row=workD[day]).value in projectperiodnamegiveninclasstt ):
        ret =  True
    if( hr > 2):
            ret = classtt.cell(column=classH[hr - 2], row=workD[day]).value in projectperiodnamegiveninclasstt or ret
    if (hr > 1):
         ret = classtt.cell(column = classH[hr - 1], row=workD[day]).value in projectperiodnamegiveninclasstt or ret
    return ret
def isGuideInBetaProjectPeriod(hr,day,guideslno):
    projectperiodnamegiveninguidestt = ["PROJECT"]
    className = ["S8 CS B","S8 CSB"]
    ret = False
    if (hr > 1):
        cellvalue1awayunder = guidestt.cell(column = classH[hr - 1], row=guideday(workD[day],guideslno)+1).value
        cellvalue1away = guidestt.cell(column = classH[hr - 1], row=guideday(workD[day],guideslno)).value
    if (hr > 2):
        cellvalue2away = guidestt.cell(column=classH[hr - 2], row=guideday(workD[day],guideslno)).value
        cellvalue2awayunder = guidestt.cell(column=classH[hr - 2], row=guideday(workD[day],guideslno)+1).value
    cellvalue = guidestt.cell(column=classH[hr], row=guideday(workD[day],guideslno)).value
    cellvalueunder = guidestt.cell(column=classH[hr], row=guideday(workD[day],guideslno)+1).value
    if((cellvalue in projectperiodnamegiveninguidestt) and ( cellvalueunder in className )):
        ret = True
    if( hr > 2):
            ret = ((( cellvalue2away in projectperiodnamegiveninguidestt) and ( cellvalue2awayunder in className )) or ret)
    if (hr > 1):
         ret = ((( cellvalue1away in projectperiodnamegiveninguidestt) and ( cellvalue1awayunder in className)) or ret)
    return ret

def twoCellsWithinPeriod(hr,day,guideslno):
    period = set()
    cellvalue1away = "XX"
    cellvalue2away = "XX"
    if (hr > 1):
        cellvalue1away = guidestt.cell(column=classH[hr - 1], row=guideday(workD[day], guideslno)).value
    if (hr > 2):
        cellvalue2away = guidestt.cell(column=classH[hr - 2], row=guideday(workD[day], guideslno)).value
    cellvalue = guidestt.cell(column=classH[hr], row=guideday(workD[day], guideslno)).value
    period = {cellvalue1away,cellvalue2away,cellvalue}
    return period

row_workD = { 3:1,5:2,7:3,9:4,12:5}
column_classH = {3:1,4:2,6:3,7:4,9:5,11:6,12:7}

long3Periods = set()
for p in ["PROJECT","MICROPROCESSOR LAB","C LAB","CP LAB","DIGITAL LAB","MP LAB","NW LAB","FOSS LAB","N/W LAB","MINI PROJECT"]:
    long3Periods.add(p)

def isFreePeriod(hr,day,guideslno=1):
    around = twoCellsWithinPeriod(column_classH[hr],row_workD[day],guideslno)
    return ( guidestt.cell(column=hr, row=guideday(day,guideslno) ).value is  None ) and ( around.intersection(long3Periods) == set() )
def guideday(day,slno=1):
    return (slno-1)*15 + day

FILENAME1 = "_class_tt_org.xlsx"
SHEETNAME1 = "Sheet1"
FILENAME2 = "_guide_tt_org.xlsx"
SHEETNAME2 = "Sheet1"

workbook1 = load_workbook(filename=FILENAME1)
workbook2 = load_workbook(filename=FILENAME2)


classtt = workbook1[SHEETNAME1]
guidestt = workbook2[SHEETNAME2]

workD = { 1:3,2:5,3:7,4:9,5:12}
classH = {1:3,2:4,3:6,4:7,5:9,6:11,7:12}
"""
Date to Begin Allocation
"""
print("Start Date")
dx = input("Day")
mx = input("Month")
yx = input("Year")


print(dx,mx,yx)
startDate = datetime.date(day=int(dx),month=int(mx),year=int(yx))

weekday = startDate.weekday() + 1
startDay = weekday
startHour = 1
endHour = 7
"""
Circular Week Days
"""
def circular_days(start):
    days = {1:1,2:2,3:3,4:4,5:5}

    if (start > 0 and start % 5 == 0):
        start = 5
    else:
        start = start % 5
    yield days[start]
    start = start + 1
    if (start > 0 and start % 5 == 0):
        start = 5
    else:
        start = start % 5
    yield days[start]
    start = start + 1
    if (start > 0 and start % 5 == 0):
        start = 5
    else:
        start = start % 5
    yield days[start]
    start = start + 1
    if (start > 0 and start % 5 == 0):
        start = 5
    else:
        start = start % 5
    yield days[start]
    start = start + 1
    if (start > 0 and start % 5 == 0):
        start = 5
    else:
        start = start % 5
    yield days[start]
    start = start + 1
days = []
for day in circular_days(startDay):
    days.append(day)
print(days)

hrs = []
"""
number of teams per project period

"""
mul = input("number of teams per project period")
multiplicity = int(mul)

for hr in range(startHour,endHour+1):
    for i in range(multiplicity):
        hrs.append(hr)
print(hrs)


"""
teams as given in S8 details file, Meenu Mathew Group Ignored while numbering and assigning 
total allocated groups 18 in number
"""
projectPeriods_notallocated = {}
MAX_TEAM_COUNT = 18
team_guide_slno = {1:25,2:19,3:5,4:13,5:8,6:19,7:12,8:16,9:9,10:9,11:6,12:16,13:4,14:4,15:23,16:17,17:29,18:29}
for i in range(10000):
    allocated = set()
    notAllocated = []
    while( len(notAllocated) < MAX_TEAM_COUNT):
        randTeam = randint(1,MAX_TEAM_COUNT)
        if( randTeam not in notAllocated):
            notAllocated.append(randTeam)
        else:
            pass
    notAllocatedOrginal = tuple(notAllocated)
    projectPeriod = 0
    allocatedPeriods = 0
    while( allocated != {1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18} ):
        for day in days:
            for hr in hrs:
                if( isProjectPeriod(hr,day) ):
                    projectPeriod = projectPeriod + 1
                    before = allocated
                    for team in notAllocated:
                        if( isGuideInBetaProjectPeriod(hr,day,team_guide_slno[team]) ):
                            """
                            allocating team with its guide having project hour in cse B
                            """
                            allocatedPeriods = allocatedPeriods + 1
                            allocated.add(team)
                            notAllocated.remove(team)
                            break
                        elif( isFreePeriod(classH[hr],workD[day],team_guide_slno[team])):
                            """
                            allpyocating as teams guide is free
                            """
                            allocatedPeriods = allocatedPeriods + 1
                            allocated.add(team)
                            notAllocated.remove(team)
                            break
                        else:
                            pass
                    after = allocated
                    if(after == before):
                        pass
    if( projectPeriod not in projectPeriods_notallocated.keys()):
        projectPeriods_notallocated[projectPeriod] = set()
        projectPeriods_notallocated[projectPeriod].add(notAllocatedOrginal )
    else:
        projectPeriods_notallocated[projectPeriod].add(notAllocatedOrginal)

final_solutions = projectPeriods_notallocated[min(projectPeriods_notallocated.keys()) ]
answer = list(final_solutions)[0]

def dthToDate(s,dth):
    ad = s
    for d in range(dth -1):
        if( (ad + datetime.timedelta(days=1)).weekday()+1 == 6 ):
            ad = ad + datetime.timedelta(days=3)
        else:
            ad = ad + datetime.timedelta(days=1)
    return ad
dthToDate(startDate,1)
# Single Iteration Selectively of Chosen
dayth = 0
allocated = set()
notAllocated = []
notAllocated = list(answer)

notAllocatedOrginal = tuple(notAllocated)
projectPeriod = 0
allocatedPeriods = 0
while (allocated != {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18}):
    for day in days:
        dayth = dayth + 1
        for hr in hrs:
            if (isProjectPeriod(hr, day)):
                projectPeriod = projectPeriod + 1
                before = allocated
                for team in notAllocated:
                    if (isGuideInBetaProjectPeriod(hr, day, team_guide_slno[team])):
                        """
                        allocating team with its guide having project hour in cse B
                        """
                        allocatedPeriods = allocatedPeriods + 1
                        allocated.add(team)
                        notAllocated.remove(team)
                        print("team:?",team,", allocated on:",hr,day,dthToDate(startDate,dayth) )
                        break
                    elif (isFreePeriod(classH[hr], workD[day], team_guide_slno[team])):
                        """
                        allpyocating as teams guide is free
                        """
                        allocatedPeriods = allocatedPeriods + 1
                        allocated.add(team)
                        notAllocated.remove(team)
                        print("team:!", team, ", allocated on:", hr, day, dthToDate(startDate,dayth)  )
                        break
                    else:
                        pass
                after = allocated
                if (after == before):
                    pass